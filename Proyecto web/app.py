from flask import Flask, render_template, request, redirect, session
from openpyxl import load_workbook
import os

app = Flask(__name__)
app.secret_key = "farmacia123"


# LEER PRODUCTOS EXCEL

def leer_productos():
    archivo = "productos.xlsx"

    if not os.path.exists(archivo):
        return []

    wb = load_workbook(archivo)
    ws = wb.active

    productos = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row or not row[0]:
            continue

        productos.append({
    "nombre": row[0] if len(row) > 0 else "",
    "precio": f"S/. {row[1]}",
    "precio_num": float(row[1]),
    "categoria": row[2] if len(row) > 2 else "",
    "imagen": row[3] if len(row) > 3 else "",
    "categoria_key": row[4] if len(row) > 4 else ""
})
    return productos

# HOME

@app.route("/")
def home():

    destacados = [
        {
            "nombre": "Ibuprofeno 600 mg",
            "precio": "S/. 3.00",
            "precio_num": 3,
            "badge": "🔥 Más vendido",
            "imagen": "https://calox.com/wp-content/uploads/2022/12/Ibuprofeno-600-mg-10-Tab-Rec.webp"
        },
        {
            "nombre": "Paracetamol Jarabe",
            "precio": "S/. 2.00",
            "precio_num": 2,
            "badge": "⭐ Recomendado",
            "imagen": "https://dcuk1cxrnzjkh.cloudfront.net/imagesproducto/015496L.jpg"
        },
        {
            "nombre": "Vitamina C",
            "precio": "S/. 5.00",
            "precio_num": 5,
            "badge": "💥 Oferta",
            "imagen": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSJRVo9kXKPePn0nrPGZSedjTDPhPsAoFHWVx2ZYGH8uVvkdkhgQxGxS0Q&s=10"
        }
    ]

    servicios = [
        {
            "nombre": "Atención farmacéutica",
            "descripcion": "Orientación básica sobre productos disponibles en farmacia.",
            "icono": "fas fa-user-md"
        },
        {
            "nombre": "Delivery de medicamentos",
            "descripcion": "Envíos rápidos de productos de salud hasta tu domicilio.",
            "icono": "fas fa-motorcycle"
        },
        {
            "nombre": "Consulta de productos",
            "descripcion": "Te ayudamos a encontrar medicamentos y vitaminas disponibles.",
            "icono": "fas fa-notes-medical"
        }
    ]

    return render_template(
        "index.html",
        destacados=destacados,
        productos=leer_productos(),
        servicios=servicios
    )


# LOGIN

@app.route("/login", methods=["POST"])
def login():

    usuario = request.form["usuario"]
    password = request.form["password"]

    if usuario == "DeYauri" and password == "022580":
        session["admin"] = True
        return redirect("/admin")

    return redirect("/")


# ADMIN

@app.route("/admin")
def admin():
    if not session.get("admin"):
        return redirect("/")

    return render_template(
        "admin.html",
        productos=leer_productos()
    )

@app.route("/agregar_producto", methods=["POST"])
def agregar_producto():

    if not session.get("admin"):
        return redirect("/")

    nombre = request.form["nombre"]
    precio = request.form["precio"]
    categoria = request.form["categoria"]
    imagen = request.form["imagen"]
    categoria_key = request.form["categoria_key"]

    wb = load_workbook("productos.xlsx")
    ws = wb.active

    ws.append([
        nombre,
        precio,
        categoria,
        imagen,
        categoria_key
    ])

    wb.save("productos.xlsx")

    return redirect("/admin")

@app.route("/eliminar_producto/<int:index>")
def eliminar_producto(index):

    if not session.get("admin"):
        return redirect("/")

    wb = load_workbook("productos.xlsx")
    ws = wb.active

    ws.delete_rows(index + 2)

    wb.save("productos.xlsx")

    return redirect("/admin")

# LOGOUT

@app.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect("/")

# RUN

if __name__ == "__main__":
    app.run(debug=True)